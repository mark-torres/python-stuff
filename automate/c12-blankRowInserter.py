#!/usr/bin/env python3
import argparse, openpyxl, sys, os
from openpyxl.cell import get_column_letter

parser = argparse.ArgumentParser(description='Inserts N blank rows at desired position in a spreadsheet')

parser.add_argument('--number','-n',
	dest='nRows',
	type=int,
	required=True,
	nargs=1,
	help="Rows to insert. Integer in the range 1 - 100")
parser.add_argument('--row','-r',
	dest='startRow',
	type=int,
	required=True,
	nargs=1,
	help="Row number where the rows will be inserted. Positive integer in the range from 1 to the number of rows in the spreadsheet.")
parser.add_argument('--file','-f',
	dest='targetFile',
	required=True,
	nargs=1,
	help="Target XLSX file")
args = parser.parse_args()

# required args
nRows = args.nRows[0]
startRow = args.startRow[0]
targetFile = args.targetFile[0]

# VALIDATE ARGUMENTS
# nRows
if nRows < 1 or nRows > 100:
	print("Number of rows to insert must be in the range 1 - 100")
	sys.exit()

# targetFile
if not os.path.exists(targetFile):
	print("Target file does not exist")
	sys.exit()
try:
	wb = openpyxl.load_workbook(targetFile)
except:
	print( "Error reading file %s" % (targetFile) )
	sys.exit()

# get highest row and column
sheet = wb.active
# maxCol = sheet.get_highest_column()
maxCol = sheet.max_column
# maxRow = sheet.get_highest_row()
maxRow = sheet.max_row

if maxRow < startRow:
	print("Start row must be less than the maximum number of rows in the spreadsheet")
	sys.exit()

# move data
lastRow = maxRow + nRows
for row in range(lastRow, lastRow - nRows - 1, -1):
	for col in range(1, maxCol + 1, 1):
		dstCell = "%s%d" % (get_column_letter(col), row)
		srcCell = "%s%d" % (get_column_letter(col), row - nRows)
		sheet[dstCell] = sheet[srcCell].value

# erase inserted rows
for row in range(startRow, startRow + nRows, 1):
	for col in range(1, maxCol + 1, 1):
		cell = "%s%d" % (get_column_letter(col), row)
		sheet[cell] = ''

wb.save(targetFile)
