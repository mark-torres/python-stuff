#!/usr/bin/env python3
import argparse, openpyxl, sys, os
from openpyxl.cell import get_column_letter
import logging
logging.basicConfig(level=logging.DEBUG, format='%(levelname)s - %(message)s')
logging.disable(logging.CRITICAL)

parser = argparse.ArgumentParser(description='Inverts columns and rows in a spreadsheet')
parser.add_argument('--file','-f',
	dest='targetFile',
	required=True,
	nargs=1,
	help="Target XLSX file")
args = parser.parse_args()

# required args
targetFile = args.targetFile[0]

# VALIDATE
# targetFile
if not os.path.exists(targetFile):
	print("Target file does not exist")
	sys.exit()
try:
	wb = openpyxl.load_workbook(targetFile)
except:
	print( "Error reading file %s" % (targetFile) )
	sys.exit()

# START
data = []
sheet = wb.active
# maxCol = sheet.get_highest_column()
maxCol = sheet.max_column
# maxRow = sheet.get_highest_row()
maxRow = sheet.max_row
# read data
for i in range(0, maxRow):
	r = i + 1
	data.append([])
	for j in range(0, maxCol):
		c = j + 1
		cell = sheet.cell( row = r, column = c)
		data[i].append(cell.value)

# invert data
invertedWb = openpyxl.Workbook()
sheet = invertedWb.active
for j in range(0, maxCol):
	for i in range(0, maxRow):
		letter = get_column_letter(i + 1)
		sheet["%s%d" % (letter, j + 1)] = data[i][j]
		logging.debug("Cell %s%d (%d,%d) = '%s'" % (letter, j + 1, i, j, data[i][j]))

# write inverted file
newFileName = os.path.basename(targetFile).split('.')[:-1]
newFileName = ".".join(newFileName)
newFileName = newFileName + "_inverted.xlsx"
invertedWb.save(os.path.join(os.path.dirname(targetFile), newFileName))
