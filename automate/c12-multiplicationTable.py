#!/usr/bin/env python3
import argparse, openpyxl, sys
from openpyxl.styles import Font, Style
from openpyxl.cell import get_column_letter

parser = argparse.ArgumentParser(description='Makes a spreadsheet with multiplication table in it')

parser.add_argument('--number','-n',
	dest='number',
	type=int,
	required=True,
	nargs=1,
	help="Maximum number for the multiplication table.")
parser.add_argument('--output','-o',
	dest='outFile',
	required=True,
	nargs=1,
	help="Output XLSX file")
args = parser.parse_args()

# required args
number = args.number[0]
outFile = args.outFile[0]

if number < 1 or number > 100:
	print('Number out of range. Please enter an integer between 1 and 100')
	sys.exit()

# create style object
boldFont = Font(bold=True)
styleBold = Style(font = boldFont)

# create woorkbook
wb = openpyxl.Workbook()
sheet = wb.active

# iterate values
maxNum = number + 1
for row in range(1, maxNum + 1):
	for col in range(1, maxNum + 1):
		# calculate cell coordinate
		colLetter = get_column_letter(col)
		cell = "%s%d" % (colLetter, row)
		# calculate cell value
		if row == 1 and col == 1:
			result = " "
		elif row == 1 and col > 1:
			result = "%d" % (col - 1)
			sheet[cell].style = styleBold
		elif row > 1 and col == 1:
			result = "%d" % (row - 1)
			sheet[cell].style = styleBold
		else:
			result = "%d" % ((col - 1) * (row - 1))
		sheet[cell] = result

# freeze pane
sheet.freeze_panes = 'B2'

if not outFile.endswith(".xlsx"):
	outFile = outFile + ".xlsx"

try:
	wb.save(outFile)
except:
	print("Error saving file.")
